import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any


def generate_attendance_summary_excel(report_data: Dict[str, Any]) -> io.BytesIO:
    """Generate an Excel workbook for attendance summary report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Summary"

    # Header Style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    # Metadata
    ws.append(["Attendance Summary Report"])
    ws.append([f"Period: {report_data['start_date']} to {report_data['end_date']}"])
    ws.append([]) # spacer

    # Table Headers
    headers = ["Employee ID", "Employee Name", "Email", "Days Worked", "Total Hours", "OT Days", "OT Hours"]
    ws.append(headers)

    for cell in ws[4]: # Headers are on row 4
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Data
    for uid, summary in report_data['summaries'].items():
        ws.append([
            summary.user_id,
            summary.user_full_name,
            summary.user_email,
            summary.days_worked,
            summary.total_hours,
            summary.overtime_days,
            summary.overtime_hours
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
